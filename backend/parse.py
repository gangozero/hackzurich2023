import csv
from openpyxl import load_workbook


def save_to_csv(filename, data):
    with open(filename, 'w') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(data)


df = load_workbook('input.xlsx', read_only=True,
                   data_only=True, keep_links=False)
sheets = df.get_sheet_names()
# print(sheets)

factories = {}
outbound = {}
inbound = {}


# outbound argentina
sheet = df[sheets[0]]

# get argentina factories
for row in sheet.iter_rows(2, sheet.max_row):
    if row[1].value not in factories:
        factories[row[1].value] = [
            row[1].value, # factory id
            row[2].value, # factory name 
            row[3].value, # lat
            row[4].value, # lon
            ]

# get argentina outbound delieveries
for row in sheet.iter_rows(2, sheet.max_row):
    if row[5].value not in outbound:
        outbound[row[5].value] = [
            row[5].value, # destination id
            row[6].value, # lat
            row[7].value, # lon
            row[18].value, # distance
            row[1].value, # factory id
        ]

# outbound mexico
sheet = df[sheets[2]]

# get mexico factories
for row in sheet.iter_rows(2, sheet.max_row):
    if row[2].value not in factories:
        factories[row[2].value] = [
            row[2].value, # factory id
            row[1].value, # factory name 
            row[3].value, # lat
            row[4].value, # lon
            ]

# get mexico outbound delieveries
# switch rows 1 and 2 since they are switched in file 
for row in sheet.iter_rows(2, sheet.max_row):
    if row[5].value not in outbound:
        outbound[row[5].value] = [
            row[5].value, # destination id
            row[6].value, # lat
            row[7].value, # lon
            row[18].value, # distance
            row[2].value, # factory id
        ]


# inbound argentina
sheet = df[sheets[1]]

for row in sheet.iter_rows(2, sheet.max_row):
    if row[1].value not in inbound and row[5].value in factories:
        inbound[row[1].value] = [
            row[1].value, # supplier id
            row[2].value, # supplier name
            row[3].value, # lat
            row[4].value, # lon
            row[18].value, # distance
            row[5].value, # factory id
        ]

# inbound mexico
sheet = df[sheets[3]]
for row in sheet.iter_rows(2, sheet.max_row):
    if row[1].value not in inbound and row[5].value in factories:
        inbound[row[1].value] = [
            row[1].value, # supplier id
            row[2].value, # supplier name
            row[3].value, # lat
            row[4].value, # lon
            row[18].value, # distance
            row[5].value, # factory id
        ]

save_to_csv('factories.csv', factories.values())
save_to_csv('outbound.csv', outbound.values())
save_to_csv('inbound.csv', inbound.values())


